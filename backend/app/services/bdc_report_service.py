"""
BDC lead export report: flexible filtering, guest QR generation, XLSX/ZIP export.
"""
from __future__ import annotations

import io
import logging
import zipfile
from dataclasses import dataclass, field
from datetime import datetime
from decimal import Decimal
from typing import Any, Dict, List, Optional, Sequence, Tuple
from uuid import UUID

import qrcode
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import and_, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.access_scope import get_accessible_dealership_ids, user_can_access_dealership
from app.core.config import settings
from app.core.timezone import utc_now
from app.models.appointment import Appointment, AppointmentStatus
from app.models.customer import Customer
from app.models.dealership import Dealership
from app.models.eligibility import EligibilityAssessment
from app.models.guest import Guest
from app.models.lead import Lead, LeadSource
from app.models.lead_stage import LeadStage
from app.models.showroom_visit import ShowroomVisit
from app.models.user import User, UserRole
from app.services.guest_service import GuestService

logger = logging.getLogger(__name__)

APPOINTMENT_FUNNELS: Dict[str, List[str]] = {
    "scheduled": [AppointmentStatus.SCHEDULED.value, AppointmentStatus.CONFIRMED.value],
    "show_up": [
        AppointmentStatus.ARRIVED.value,
        AppointmentStatus.IN_SHOWROOM.value,
        AppointmentStatus.IN_PROGRESS.value,
        AppointmentStatus.COMPLETED.value,
        AppointmentStatus.SOLD.value,
    ],
    "completed": [AppointmentStatus.COMPLETED.value],
    "sold": [AppointmentStatus.SOLD.value],
    "no_show": [AppointmentStatus.NO_SHOW.value],
    "cancelled": [AppointmentStatus.CANCELLED.value, AppointmentStatus.RESCHEDULED.value],
}

APPOINTMENT_FUNNEL_LABELS: Dict[str, str] = {
    "scheduled": "Scheduled / Confirmed",
    "show_up": "Show up (arrived+)",
    "completed": "Completed",
    "sold": "Sold (appointment)",
    "no_show": "No show",
    "cancelled": "Cancelled / Rescheduled",
}

LEAD_SOURCE_LABELS: Dict[str, str] = {
    "manual": "Manual",
    "website": "Website",
    "google_sheets": "Google Sheets",
    "meta_ads": "Meta Ads",
    "referral": "Referral",
    "walk_in": "Walk-in",
    "whatsapp_inbound": "WhatsApp",
}

PDF_TABLE_HEADERS = [
    "Lead",
    "Phone",
    "Stage",
    "Dealership",
    "BDC",
    "Sales",
    "Status",
    "Appointment",
    "Guest Trust",
    "QR",
    "Auto",
]

# Landscape A4 usable width with 10mm side margins
PDF_CONTENT_WIDTH_MM = 277

EXPORT_HEADERS = [
    "Lead ID",
    "Full Name",
    "Email",
    "Phone",
    "Stage",
    "Source",
    "Lead Created",
    "Dealership",
    "BDC Agent",
    "Salesperson",
    "Active",
    "Converted At",
    "Latest Appt Status",
    "Latest Appt Date",
    "Appt Count",
    "Showroom Check-in",
    "Guest Trust Score",
    "Guest QR URL",
    "Guest Auto-Generated",
    "Guest QR",
]


@dataclass
class BdcReportFilters:
    dealership_id: Optional[UUID] = None
    all_dealerships: bool = False
    bdc_agent_id: Optional[UUID] = None
    assigned_to: Optional[UUID] = None
    stage_id: Optional[UUID] = None
    source: Optional[str] = None
    is_active: Optional[bool] = None
    search: Optional[str] = None
    lead_date_from: Optional[datetime] = None
    lead_date_to: Optional[datetime] = None
    sold_date_from: Optional[datetime] = None
    sold_date_to: Optional[datetime] = None
    appointment_date_from: Optional[datetime] = None
    appointment_date_to: Optional[datetime] = None
    appointment_statuses: List[str] = field(default_factory=list)
    appointment_funnel: Optional[str] = None
    has_appointment: Optional[bool] = None
    sold_only: bool = False
    converted_only: bool = False


@dataclass
class BdcReportMeta:
    generated_at: datetime
    generated_by: str
    total_leads: int
    scope_label: str
    filter_items: List[Tuple[str, str]] = field(default_factory=list)


@dataclass
class BdcReportRow:
    lead_id: str
    full_name: str
    email: str
    phone: str
    stage: str
    source: str
    lead_created: str
    dealership: str
    bdc_agent: str
    salesperson: str
    is_active: bool
    converted_at: str
    latest_appt_status: str
    latest_appt_date: str
    appt_count: int
    showroom_check_in: str
    lead_trust_score: Optional[float]
    guest_trust_score: Optional[float]
    guest_qr_url: str
    guest_auto_generated: bool
    share_token: Optional[str] = None


class BdcReportService:
    @staticmethod
    async def resolve_dealership_ids(
        db: AsyncSession,
        current_user: User,
        dealership_id: Optional[UUID],
        all_dealerships: bool = False,
    ) -> List[UUID]:
        """Return the dealership scope for the report (one or many)."""
        accessible = await get_accessible_dealership_ids(db, current_user)

        if all_dealerships:
            if current_user.role == UserRole.SUPER_ADMIN:
                res = await db.execute(select(Dealership.id))
                ids = [r[0] for r in res.all()]
                if not ids:
                    raise ValueError("No dealerships found")
                return ids
            if current_user.role == UserRole.BDC and accessible:
                return list(accessible)
            if current_user.dealership_id:
                return [current_user.dealership_id]
            raise ValueError("No dealerships in scope")

        if dealership_id:
            if current_user.role == UserRole.SUPER_ADMIN:
                return [dealership_id]
            if accessible and dealership_id not in accessible:
                if current_user.dealership_id != dealership_id:
                    raise PermissionError("Not authorized for this dealership")
            elif (
                current_user.role != UserRole.SUPER_ADMIN
                and current_user.dealership_id
                and current_user.dealership_id != dealership_id
                and current_user.role != UserRole.BDC
            ):
                raise PermissionError("Not authorized for this dealership")
            return [dealership_id]

        if current_user.role == UserRole.BDC and accessible:
            if len(accessible) == 1:
                return list(accessible)
            raise ValueError("Select a dealership or choose All dealerships")

        if current_user.dealership_id:
            return [current_user.dealership_id]

        raise ValueError("Dealership context required")

    @staticmethod
    def _appointment_status_filter(
        filters: BdcReportFilters,
    ) -> Optional[List[str]]:
        statuses: List[str] = []
        if filters.appointment_funnel and filters.appointment_funnel in APPOINTMENT_FUNNELS:
            statuses.extend(APPOINTMENT_FUNNELS[filters.appointment_funnel])
        if filters.appointment_statuses:
            statuses.extend(filters.appointment_statuses)
        return list(dict.fromkeys(statuses)) or None

    @staticmethod
    async def build_leads_query(
        db: AsyncSession,
        current_user: User,
        filters: BdcReportFilters,
        dealership_ids: List[UUID],
    ):
        query = select(Lead).where(Lead.dealership_id.in_(dealership_ids))

        if filters.bdc_agent_id:
            query = query.where(Lead.bdc_assigned_to_id == filters.bdc_agent_id)
        if filters.assigned_to:
            query = query.where(Lead.assigned_to == filters.assigned_to)
        if filters.stage_id:
            query = query.where(Lead.stage_id == filters.stage_id)
        if filters.source:
            try:
                query = query.where(Lead.source == LeadSource(filters.source))
            except ValueError:
                query = query.where(Lead.id.is_(None))
        if filters.is_active is not None:
            query = query.where(Lead.is_active == filters.is_active)
        if filters.sold_only or filters.converted_only:
            query = query.where(Lead.outcome == "converted")
        if filters.lead_date_from:
            query = query.where(Lead.created_at >= filters.lead_date_from)
        if filters.lead_date_to:
            query = query.where(Lead.created_at <= filters.lead_date_to)
        if filters.sold_date_from:
            query = query.where(
                or_(
                    Lead.converted_at >= filters.sold_date_from,
                    and_(Lead.converted_at.is_(None), Lead.closed_at >= filters.sold_date_from),
                )
            )
        if filters.sold_date_to:
            query = query.where(
                or_(
                    Lead.converted_at <= filters.sold_date_to,
                    and_(Lead.converted_at.is_(None), Lead.closed_at <= filters.sold_date_to),
                )
            )
        if filters.search:
            query = query.join(Customer, Lead.customer_id == Customer.id)
            full_name = func.concat(Customer.first_name, " ", func.coalesce(Customer.last_name, ""))
            query = query.where(
                or_(
                    Customer.first_name.ilike(f"%{filters.search}%"),
                    Customer.last_name.ilike(f"%{filters.search}%"),
                    full_name.ilike(f"%{filters.search}%"),
                    Customer.email.ilike(f"%{filters.search}%"),
                    Customer.phone.ilike(f"%{filters.search}%"),
                )
            )

        statuses = BdcReportService._appointment_status_filter(filters)
        appt_conditions = [Appointment.lead_id == Lead.id, Appointment.lead_id.isnot(None)]
        if statuses:
            appt_conditions.append(Appointment.status.in_(statuses))
        if filters.appointment_date_from:
            appt_conditions.append(Appointment.scheduled_at >= filters.appointment_date_from)
        if filters.appointment_date_to:
            appt_conditions.append(Appointment.scheduled_at <= filters.appointment_date_to)

        if filters.has_appointment is True or statuses or filters.appointment_date_from or filters.appointment_date_to:
            appt_exists = select(Appointment.id).where(and_(*appt_conditions)).exists()
            query = query.where(appt_exists)
        elif filters.has_appointment is False:
            appt_exists = select(Appointment.id).where(Appointment.lead_id == Lead.id).exists()
            query = query.where(~appt_exists)

        return query.order_by(Lead.created_at.desc())

    @staticmethod
    async def fetch_rows(
        db: AsyncSession,
        current_user: User,
        filters: BdcReportFilters,
        *,
        limit: Optional[int] = None,
        ensure_guests: bool = True,
    ) -> Tuple[List[BdcReportRow], int]:
        dealership_ids = await BdcReportService.resolve_dealership_ids(
            db, current_user, filters.dealership_id, filters.all_dealerships
        )
        base_query = await BdcReportService.build_leads_query(
            db, current_user, filters, dealership_ids
        )

        count_result = await db.execute(select(func.count()).select_from(base_query.subquery()))
        total = count_result.scalar() or 0

        query = base_query
        if limit:
            query = query.limit(limit)
        result = await db.execute(query)
        leads = result.scalars().all()
        if not leads:
            return [], total

        lead_ids = [l.id for l in leads]
        cust_ids = {l.customer_id for l in leads if l.customer_id}
        stage_ids = {l.stage_id for l in leads if l.stage_id}
        user_ids = set()
        for l in leads:
            if l.assigned_to:
                user_ids.add(l.assigned_to)
            if l.bdc_assigned_to_id:
                user_ids.add(l.bdc_assigned_to_id)

        customers: Dict[UUID, Customer] = {}
        if cust_ids:
            cr = await db.execute(select(Customer).where(Customer.id.in_(cust_ids)))
            customers = {c.id: c for c in cr.scalars().all()}

        stages: Dict[UUID, LeadStage] = {}
        if stage_ids:
            sr = await db.execute(select(LeadStage).where(LeadStage.id.in_(stage_ids)))
            stages = {s.id: s for s in sr.scalars().all()}

        users: Dict[UUID, User] = {}
        if user_ids:
            ur = await db.execute(select(User).where(User.id.in_(user_ids)))
            users = {u.id: u for u in ur.scalars().all()}

        dealership_names: Dict[UUID, str] = {}
        if dealership_ids:
            dr = await db.execute(select(Dealership).where(Dealership.id.in_(dealership_ids)))
            for d in dr.scalars().all():
                dealership_names[d.id] = d.name or ""

        appt_result = await db.execute(
            select(Appointment)
            .where(Appointment.lead_id.in_(lead_ids))
            .order_by(Appointment.scheduled_at.desc())
        )
        appts_by_lead: Dict[UUID, List[Appointment]] = {}
        for appt in appt_result.scalars().all():
            if appt.lead_id:
                appts_by_lead.setdefault(appt.lead_id, []).append(appt)

        checkin_result = await db.execute(
            select(ShowroomVisit.lead_id).where(ShowroomVisit.lead_id.in_(lead_ids)).distinct()
        )
        checked_in_leads = {r[0] for r in checkin_result.all()}

        guest_result = await db.execute(select(Guest).where(Guest.lead_id.in_(lead_ids)))
        guests_by_lead: Dict[UUID, Guest] = {}
        for g in guest_result.scalars().all():
            if g.lead_id and g.lead_id not in guests_by_lead:
                guests_by_lead[g.lead_id] = g

        lead_scores: Dict[UUID, float] = {}
        guest_scores: Dict[UUID, float] = {}
        guest_ids = [g.id for g in guests_by_lead.values()]
        if lead_ids:
            lr = await db.execute(
                select(EligibilityAssessment).where(
                    EligibilityAssessment.entity_type == "lead",
                    EligibilityAssessment.entity_id.in_(lead_ids),
                )
            )
            for a in lr.scalars().all():
                lead_scores[a.entity_id] = float(a.total_score)
        if guest_ids:
            gr = await db.execute(
                select(EligibilityAssessment).where(
                    EligibilityAssessment.entity_type == "guest",
                    EligibilityAssessment.entity_id.in_(guest_ids),
                )
            )
            for a in gr.scalars().all():
                guest_scores[a.entity_id] = float(a.total_score)

        frontend = settings.frontend_url.rstrip("/")
        rows: List[BdcReportRow] = []

        for lead in leads:
            customer = customers.get(lead.customer_id)
            full_name = ""
            if customer:
                full_name = f"{customer.first_name or ''} {customer.last_name or ''}".strip()
            stage = stages.get(lead.stage_id)
            bdc_user = users.get(lead.bdc_assigned_to_id) if lead.bdc_assigned_to_id else None
            sp_user = users.get(lead.assigned_to) if lead.assigned_to else None

            lead_appts = appts_by_lead.get(lead.id, [])
            latest_appt = lead_appts[0] if lead_appts else None

            guest = guests_by_lead.get(lead.id)
            auto_generated = False
            if ensure_guests:
                if not guest:
                    guest = await GuestService.ensure_for_lead(
                        db,
                        lead.id,
                        dealership_id=lead.dealership_id,
                    )
                    guests_by_lead[lead.id] = guest
                    auto_generated = True
                elif not guest.share_token:
                    GuestService.ensure_share_token(guest)

            share_url = ""
            share_token = None
            guest_trust: Optional[float] = None
            if guest and guest.share_token and not guest.share_revoked:
                share_token = guest.share_token
                share_url = f"{frontend}/g/{guest.share_token}"
                guest_trust = guest_scores.get(guest.id)

            source_display = (lead.meta_data or {}).get("source_display") or (
                lead.source.value if lead.source else ""
            )

            rows.append(
                BdcReportRow(
                    lead_id=str(lead.id),
                    full_name=full_name or "Unknown",
                    email=(customer.email if customer else "") or "",
                    phone=(customer.phone if customer else "") or "",
                    stage=stage.display_name if stage else "",
                    source=source_display,
                    lead_created=lead.created_at.isoformat() if lead.created_at else "",
                    dealership=dealership_names.get(lead.dealership_id, "") if lead.dealership_id else "",
                    bdc_agent=bdc_user.full_name if bdc_user else "",
                    salesperson=sp_user.full_name if sp_user else "",
                    is_active=bool(lead.is_active),
                    converted_at=(
                        (lead.converted_at or lead.closed_at).isoformat()
                        if (lead.converted_at or lead.closed_at)
                        else ""
                    ),
                    latest_appt_status=(
                        latest_appt.status.replace("_", " ").title() if latest_appt else ""
                    ),
                    latest_appt_date=(
                        latest_appt.scheduled_at.isoformat() if latest_appt and latest_appt.scheduled_at else ""
                    ),
                    appt_count=len(lead_appts),
                    showroom_check_in="Yes" if lead.id in checked_in_leads else "No",
                    lead_trust_score=lead_scores.get(lead.id),
                    guest_trust_score=guest_trust,
                    guest_qr_url=share_url,
                    guest_auto_generated=auto_generated,
                    share_token=share_token,
                )
            )

        if ensure_guests:
            await db.commit()

        return rows, total

    @staticmethod
    def _format_display_datetime(dt: datetime) -> str:
        if dt.tzinfo is not None:
            dt = dt.replace(tzinfo=None)
        hour = dt.strftime("%I").lstrip("0") or "12"
        return f"{dt.strftime('%A, %b %d, %Y')} at {hour}:{dt.strftime('%M %p')}"

    @staticmethod
    def _format_filter_date(dt: Optional[datetime]) -> str:
        if not dt:
            return ""
        if dt.tzinfo is not None:
            dt = dt.replace(tzinfo=None)
        return dt.strftime("%b %d, %Y")

    @staticmethod
    def _format_filter_date_range(
        date_from: Optional[datetime],
        date_to: Optional[datetime],
    ) -> Optional[str]:
        start = BdcReportService._format_filter_date(date_from)
        end = BdcReportService._format_filter_date(date_to)
        if start and end:
            return f"{start} – {end}"
        if start:
            return f"From {start}"
        if end:
            return f"Through {end}"
        return None

    @staticmethod
    async def build_report_meta(
        db: AsyncSession,
        current_user: User,
        filters: BdcReportFilters,
        dealership_ids: List[UUID],
        total_leads: int,
    ) -> BdcReportMeta:
        dealership_names: List[str] = []
        if dealership_ids:
            dr = await db.execute(select(Dealership).where(Dealership.id.in_(dealership_ids)))
            dealership_names = sorted({d.name for d in dr.scalars().all() if d.name})

        if filters.all_dealerships:
            if len(dealership_names) == 1:
                scope_label = dealership_names[0]
            elif dealership_names:
                scope_label = f"All Dealerships ({len(dealership_names)})"
            else:
                scope_label = "All Dealerships"
        elif filters.dealership_id and dealership_names:
            scope_label = dealership_names[0]
        elif filters.dealership_id:
            scope_label = str(filters.dealership_id)[:8]
        else:
            scope_label = dealership_names[0] if len(dealership_names) == 1 else "Selected Dealerships"

        filter_items: List[Tuple[str, str]] = [("Dealership scope", scope_label)]

        user_ids: List[UUID] = []
        if filters.bdc_agent_id:
            user_ids.append(filters.bdc_agent_id)
        if filters.assigned_to:
            user_ids.append(filters.assigned_to)
        users: Dict[UUID, User] = {}
        if user_ids:
            ur = await db.execute(select(User).where(User.id.in_(user_ids)))
            users = {u.id: u for u in ur.scalars().all()}

        if filters.bdc_agent_id:
            agent = users.get(filters.bdc_agent_id)
            filter_items.append(("BDC agent", agent.full_name if agent else str(filters.bdc_agent_id)[:8]))
        if filters.assigned_to:
            sp = users.get(filters.assigned_to)
            filter_items.append(("Salesperson", sp.full_name if sp else str(filters.assigned_to)[:8]))
        if filters.stage_id:
            sr = await db.execute(select(LeadStage).where(LeadStage.id == filters.stage_id))
            stage = sr.scalar_one_or_none()
            filter_items.append(("Lead stage", stage.display_name if stage else str(filters.stage_id)[:8]))
        if filters.source:
            filter_items.append(
                ("Lead source", LEAD_SOURCE_LABELS.get(filters.source, filters.source.replace("_", " ").title()))
            )
        if filters.search:
            filter_items.append(("Search", filters.search))
        if filters.is_active is True:
            filter_items.append(("Lead status", "Active only"))
        elif filters.is_active is False:
            filter_items.append(("Lead status", "Inactive only"))

        lead_range = BdcReportService._format_filter_date_range(
            filters.lead_date_from, filters.lead_date_to
        )
        if lead_range:
            filter_items.append(("Lead created", lead_range))

        sold_range = BdcReportService._format_filter_date_range(
            filters.sold_date_from, filters.sold_date_to
        )
        if sold_range:
            filter_items.append(("Sold / converted date", sold_range))

        appt_range = BdcReportService._format_filter_date_range(
            filters.appointment_date_from, filters.appointment_date_to
        )
        if appt_range:
            filter_items.append(("Appointment date", appt_range))

        if filters.appointment_funnel:
            filter_items.append(
                (
                    "Appointment funnel",
                    APPOINTMENT_FUNNEL_LABELS.get(
                        filters.appointment_funnel,
                        filters.appointment_funnel.replace("_", " ").title(),
                    ),
                )
            )
        if filters.appointment_statuses:
            labels = [s.replace("_", " ").title() for s in filters.appointment_statuses]
            filter_items.append(("Appointment statuses", ", ".join(labels)))
        if filters.has_appointment is True:
            filter_items.append(("Has appointment", "Yes"))
        elif filters.has_appointment is False:
            filter_items.append(("Has appointment", "No"))
        if filters.sold_only:
            filter_items.append(("Outcome", "Sold only"))
        elif filters.converted_only:
            filter_items.append(("Outcome", "Converted only"))

        extra_filters = [item for item in filter_items if item[0] != "Dealership scope"]
        if not extra_filters:
            filter_items.append(("Additional filters", "None (all leads in scope)"))

        return BdcReportMeta(
            generated_at=utc_now(),
            generated_by=current_user.full_name or current_user.email or "Unknown",
            total_leads=total_leads,
            scope_label=scope_label,
            filter_items=filter_items,
        )

    @staticmethod
    def build_export_filename(meta: BdcReportMeta, ext: str) -> str:
        scope = BdcReportService._sanitize_filename_part(meta.scope_label, 45)
        date_part = meta.generated_at.strftime("%Y-%m-%d")
        return f"BDC Export - {scope} - {date_part}.{ext}"

    @staticmethod
    def _format_trust_score(row: BdcReportRow) -> str:
        if row.guest_trust_score is not None:
            return f"{int(round(row.guest_trust_score))}"
        return ""

    @staticmethod
    def _qr_png_bytes(url: str) -> bytes:
        qr = qrcode.QRCode(version=1, box_size=4, border=2)
        qr.add_data(url)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        return buf.getvalue()

    @staticmethod
    def _sanitize_filename_part(text: str, max_len: int = 60) -> str:
        """Make a string safe for use in archive file names."""
        if not text:
            return ""
        cleaned = text
        for ch in '<>:"/\\|?*\n\r\t':
            cleaned = cleaned.replace(ch, "-")
        cleaned = " ".join(cleaned.split()).strip(" .")
        return cleaned[:max_len] or "Lead"

    @staticmethod
    def _format_appt_for_filename(iso_date: str) -> str:
        """Human-readable appointment label for QR file names, e.g. Monday - Jul 5, 2026 - 2:30 PM."""
        if not iso_date or not iso_date.strip():
            return "No Appointment"
        try:
            raw = iso_date.strip()
            if raw.endswith("Z"):
                raw = raw[:-1] + "+00:00"
            dt = datetime.fromisoformat(raw)
            if dt.tzinfo is not None:
                dt = dt.replace(tzinfo=None)
            weekday = dt.strftime("%A")
            date_part = f"{dt.strftime('%b')} {dt.day}, {dt.strftime('%Y')}"
            hour = dt.strftime("%I").lstrip("0") or "12"
            time_part = f"{hour}:{dt.strftime('%M')} {dt.strftime('%p')}"
            return f"{weekday} - {date_part} - {time_part}"
        except (ValueError, TypeError):
            return "Appointment"

    @staticmethod
    def _qr_png_filename(row: BdcReportRow, used_names: set[str]) -> str:
        """Build a unique PNG file name: Lead Name - Monday - Jul 5, 2026 - 2:30 PM.png"""
        lead_name = BdcReportService._sanitize_filename_part(row.full_name, 50) or row.lead_id[:8]
        appt_label = BdcReportService._sanitize_filename_part(
            BdcReportService._format_appt_for_filename(row.latest_appt_date),
            80,
        )
        base = f"{lead_name} - {appt_label}.png"
        if base not in used_names:
            used_names.add(base)
            return base
        stem = f"{lead_name} - {appt_label}"
        n = 2
        while True:
            candidate = f"{stem} ({n}).png"
            if candidate not in used_names:
                used_names.add(candidate)
                return candidate
            n += 1

    @staticmethod
    def build_xlsx(rows: Sequence[BdcReportRow], meta: Optional[BdcReportMeta] = None) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "BDC Export"

        highlight_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        brand_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        brand_font = Font(bold=True, size=16, color="FFFFFF")
        title_font = Font(bold=True, size=14)
        label_font = Font(bold=True)

        col_count = len(EXPORT_HEADERS)
        current_row = 1

        if meta:
            brand_cell = ws.cell(
                row=current_row,
                column=1,
                value=f"{settings.app_name}  —  BDC Export Report",
            )
            brand_cell.font = brand_font
            brand_cell.fill = brand_fill
            ws.merge_cells(
                start_row=current_row,
                start_column=1,
                end_row=current_row,
                end_column=col_count,
            )
            ws.row_dimensions[current_row].height = 26
            for col_idx in range(1, col_count + 1):
                ws.cell(row=current_row, column=col_idx).fill = brand_fill
            current_row += 1

            summary = (
                f"Generated: {BdcReportService._format_display_datetime(meta.generated_at)}"
                f"  |  By: {meta.generated_by}"
                f"  |  {meta.total_leads} lead{'s' if meta.total_leads != 1 else ''}"
            )
            ws.cell(row=current_row, column=1, value=summary)
            ws.merge_cells(
                start_row=current_row,
                start_column=1,
                end_row=current_row,
                end_column=col_count,
            )
            current_row += 2

            ws.cell(row=current_row, column=1, value="Applied Filters").font = Font(bold=True, size=11)
            ws.merge_cells(
                start_row=current_row,
                start_column=1,
                end_row=current_row,
                end_column=col_count,
            )
            current_row += 1

            for label, value in meta.filter_items:
                ws.cell(row=current_row, column=1, value=label).font = label_font
                ws.cell(row=current_row, column=2, value=value)
                ws.merge_cells(
                    start_row=current_row,
                    start_column=2,
                    end_row=current_row,
                    end_column=col_count,
                )
                current_row += 1

            current_row += 1

        header_row = current_row
        for col_idx, header in enumerate(EXPORT_HEADERS, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill

        qr_url_col = EXPORT_HEADERS.index("Guest QR URL") + 1
        ws.column_dimensions[get_column_letter(qr_url_col)].width = 40
        ws.column_dimensions[get_column_letter(len(EXPORT_HEADERS))].width = 14

        for row_idx, row in enumerate(rows, start=header_row + 1):
            appt_display = (
                BdcReportService._format_appt_for_filename(row.latest_appt_date)
                if row.latest_appt_date
                else ""
            )
            values = [
                row.lead_id,
                row.full_name,
                row.email,
                row.phone,
                row.stage,
                row.source,
                row.lead_created,
                row.dealership,
                row.bdc_agent,
                row.salesperson,
                "Yes" if row.is_active else "No",
                row.converted_at,
                row.latest_appt_status,
                appt_display or row.latest_appt_date,
                row.appt_count,
                row.showroom_check_in,
                row.guest_trust_score if row.guest_trust_score is not None else "",
                row.guest_qr_url,
                "Yes" if row.guest_auto_generated else "No",
                "",
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row.guest_auto_generated:
                    cell.fill = highlight_fill

            if row.guest_qr_url:
                try:
                    png = BdcReportService._qr_png_bytes(row.guest_qr_url)
                    img_buf = io.BytesIO(png)
                    xl_img = XLImage(img_buf)
                    xl_img.width = 72
                    xl_img.height = 72
                    qr_col = EXPORT_HEADERS.index("Guest QR") + 1
                    ws.row_dimensions[row_idx].height = 58
                    ws.add_image(xl_img, f"{get_column_letter(qr_col)}{row_idx}")
                except Exception:
                    logger.exception("Failed to embed QR for lead %s", row.lead_id)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    @staticmethod
    def _pdf_escape(text: str) -> str:
        return (
            (text or "")
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
        )

    @staticmethod
    def _pdf_cell(text: str, style: Any) -> Any:
        from reportlab.platypus import Paragraph

        safe = BdcReportService._pdf_escape(text.strip() if text else "")
        return Paragraph(safe or "—", style)

    @staticmethod
    def _format_appt_for_pdf(iso_date: str) -> str:
        """Compact two-line appointment label for PDF cells."""
        if not iso_date or not iso_date.strip():
            return "—"
        try:
            raw = iso_date.strip()
            if raw.endswith("Z"):
                raw = raw[:-1] + "+00:00"
            dt = datetime.fromisoformat(raw)
            if dt.tzinfo is not None:
                dt = dt.replace(tzinfo=None)
            date_line = f"{dt.strftime('%a')}, {dt.strftime('%b')} {dt.day}, {dt.strftime('%Y')}"
            hour = dt.strftime("%I").lstrip("0") or "12"
            time_line = f"{hour}:{dt.strftime('%M %p')}"
            return f"{date_line}<br/>{time_line}"
        except (ValueError, TypeError):
            return "—"

    @staticmethod
    def _format_trust_score_pdf(row: BdcReportRow) -> str:
        if row.guest_trust_score is not None:
            return f"{int(round(row.guest_trust_score))}"
        return "—"

    @staticmethod
    def _pdf_column_widths() -> List[float]:
        from reportlab.lib.units import mm

        # Must sum to PDF_CONTENT_WIDTH_MM
        ratios = [14, 10, 9, 12, 9, 9, 8, 13, 7, 6, 5]
        total = sum(ratios)
        return [(PDF_CONTENT_WIDTH_MM * r / total) * mm for r in ratios]

    @staticmethod
    def build_zip(rows: Sequence[BdcReportRow], meta: Optional[BdcReportMeta] = None) -> bytes:
        xlsx_bytes = BdcReportService.build_xlsx(rows, meta)
        pdf_bytes = BdcReportService.build_pdf(rows, meta)
        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("BDC Export Report.pdf", pdf_bytes)
            zf.writestr("BDC Export Report.xlsx", xlsx_bytes)
            used_qr_names: set[str] = set()
            for row in rows:
                if not row.guest_qr_url:
                    continue
                png_name = BdcReportService._qr_png_filename(row, used_qr_names)
                zf.writestr(f"qr-codes/{png_name}", BdcReportService._qr_png_bytes(row.guest_qr_url))
            readme_lines = [
                f"{settings.app_name} — BDC Export Report",
                "=" * (len(settings.app_name) + 22),
                "",
            ]
            if meta:
                readme_lines.extend(
                    [
                        f"Generated: {BdcReportService._format_display_datetime(meta.generated_at)}",
                        f"Exported by: {meta.generated_by}",
                        f"Total leads: {meta.total_leads}",
                        "",
                        "Applied Filters",
                        "---------------",
                    ]
                )
                for label, value in meta.filter_items:
                    readme_lines.append(f"  {label}: {value}")
                readme_lines.append("")
            readme_lines.extend(
                [
                    "Contents",
                    "--------",
                    "BDC Export Report.pdf — printable report with QR codes in each row.",
                    "BDC Export Report.xlsx — spreadsheet with embedded QR images.",
                    "qr-codes/ — individual PNG QR images per lead.",
                    "  File names: {Lead Name} - {Weekday} - {Mon D, YYYY} - {H:MM AM/PM}.png",
                    "",
                    "Rows highlighted in yellow had guest profiles auto-generated during export.",
                ]
            )
            zf.writestr("README.txt", "\n".join(readme_lines))
        return zip_buf.getvalue()

    @staticmethod
    def build_pdf(rows: Sequence[BdcReportRow], meta: Optional[BdcReportMeta] = None) -> bytes:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import Image as RLImage, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

        page_w, _page_h = landscape(A4)
        content_w = PDF_CONTENT_WIDTH_MM * mm
        side_margin = (page_w - content_w) / 2

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf,
            pagesize=landscape(A4),
            leftMargin=side_margin,
            rightMargin=side_margin,
            topMargin=8 * mm,
            bottomMargin=14 * mm,
        )

        styles = getSampleStyleSheet()
        banner_title = ParagraphStyle(
            "BannerTitle",
            parent=styles["Heading1"],
            fontName="Helvetica-Bold",
            fontSize=18,
            leading=22,
            textColor=colors.white,
            spaceAfter=2,
        )
        banner_sub = ParagraphStyle(
            "BannerSub",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=8.5,
            leading=12,
            textColor=colors.HexColor("#DBEAFE"),
        )
        brand_style = ParagraphStyle(
            "BrandMark",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=15,
            leading=18,
            textColor=colors.white,
            alignment=2,  # right
        )
        brand_tag_style = ParagraphStyle(
            "BrandTag",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=7,
            leading=10,
            textColor=colors.HexColor("#BFDBFE"),
            alignment=2,  # right
        )
        panel_title = ParagraphStyle(
            "PanelTitle",
            parent=styles["Heading2"],
            fontName="Helvetica-Bold",
            fontSize=9,
            leading=12,
            textColor=colors.HexColor("#1E293B"),
            spaceAfter=6,
        )
        filter_label = ParagraphStyle(
            "FilterLabel",
            fontName="Helvetica-Bold",
            fontSize=7.5,
            leading=10,
            textColor=colors.HexColor("#64748B"),
        )
        filter_value = ParagraphStyle(
            "FilterValue",
            fontName="Helvetica",
            fontSize=7.5,
            leading=10,
            textColor=colors.HexColor("#0F172A"),
        )
        note_style = ParagraphStyle(
            "ReportNote",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=7,
            leading=10,
            textColor=colors.HexColor("#92400E"),
        )
        header_cell = ParagraphStyle(
            "HeaderCell",
            fontName="Helvetica-Bold",
            fontSize=7.5,
            leading=9,
            textColor=colors.white,
        )
        body_cell = ParagraphStyle(
            "BodyCell",
            fontName="Helvetica",
            fontSize=7,
            leading=9,
            textColor=colors.HexColor("#1E293B"),
        )
        body_cell_center = ParagraphStyle(
            "BodyCellCenter",
            parent=body_cell,
            alignment=1,
        )

        elements: List[Any] = []

        if meta:
            meta_line = (
                f"Generated {BdcReportService._format_display_datetime(meta.generated_at)}"
                f"  ·  Exported by <b>{BdcReportService._pdf_escape(meta.generated_by)}</b>"
                f"  ·  <b>{meta.total_leads}</b> lead{'s' if meta.total_leads != 1 else ''}"
            )
            left_cell = [
                Paragraph("BDC Export Report", banner_title),
                Paragraph(meta_line, banner_sub),
            ]
            right_cell = [
                Paragraph(BdcReportService._pdf_escape(settings.app_name), brand_style),
                Paragraph("Automotive CRM", brand_tag_style),
            ]
            banner = Table(
                [[left_cell, right_cell]],
                colWidths=[content_w * 0.68, content_w * 0.32],
            )
            banner.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#1E3A8A")),
                        ("LEFTPADDING", (0, 0), (-1, -1), 14),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 14),
                        ("TOPPADDING", (0, 0), (-1, -1), 12),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#1E40AF")),
                    ]
                )
            )
            elements.append(banner)
            elements.append(Spacer(1, 4 * mm))

            filter_rows = [
                [
                    BdcReportService._pdf_cell(label, filter_label),
                    BdcReportService._pdf_cell(value, filter_value),
                ]
                for label, value in meta.filter_items
            ]
            filter_inner = Table(
                filter_rows,
                colWidths=[38 * mm, content_w - 38 * mm - 20 * mm],
            )
            filter_inner.setStyle(
                TableStyle(
                    [
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("TOPPADDING", (0, 0), (-1, -1), 3),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                        ("LINEBELOW", (0, 0), (-1, -2), 0.25, colors.HexColor("#E2E8F0")),
                    ]
                )
            )
            filter_panel = Table(
                [[Paragraph("Applied Filters", panel_title)], [filter_inner]],
                colWidths=[content_w],
            )
            filter_panel.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
                        ("BOX", (0, 0), (-1, -1), 0.75, colors.HexColor("#CBD5E1")),
                        ("LEFTPADDING", (0, 0), (-1, -1), 10),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                        ("TOPPADDING", (0, 0), (0, 0), 8),
                        ("BOTTOMPADDING", (0, -1), (-1, -1), 10),
                    ]
                )
            )
            elements.append(filter_panel)
            elements.append(Spacer(1, 3 * mm))

            note_box = Table(
                [[Paragraph("Yellow rows = guest profiles auto-generated during this export.", note_style)]],
                colWidths=[content_w],
            )
            note_box.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#FFFBEB")),
                        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#FDE68A")),
                        ("LEFTPADDING", (0, 0), (-1, -1), 10),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                        ("TOPPADDING", (0, 0), (-1, -1), 6),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                    ]
                )
            )
            elements.append(note_box)
            elements.append(Spacer(1, 5 * mm))

        col_widths = BdcReportService._pdf_column_widths()
        table_data: List[List[Any]] = [
            [Paragraph(BdcReportService._pdf_escape(h), header_cell) for h in PDF_TABLE_HEADERS]
        ]
        highlight_rows: List[int] = []
        qr_col_idx = PDF_TABLE_HEADERS.index("QR")

        for idx, row in enumerate(rows):
            appt_display = (
                BdcReportService._format_appt_for_pdf(row.latest_appt_date)
                if row.latest_appt_date
                else "—"
            )
            trust_display = BdcReportService._format_trust_score_pdf(row)

            qr_cell: Any = BdcReportService._pdf_cell("", body_cell_center)
            if row.guest_qr_url:
                try:
                    png = BdcReportService._qr_png_bytes(row.guest_qr_url)
                    qr_cell = RLImage(io.BytesIO(png), width=11 * mm, height=11 * mm)
                except Exception:
                    qr_cell = BdcReportService._pdf_cell("—", body_cell_center)
                    logger.exception("Failed to embed QR in PDF for lead %s", row.lead_id)

            if row.guest_auto_generated:
                highlight_rows.append(idx + 1)

            table_data.append([
                BdcReportService._pdf_cell(row.full_name, body_cell),
                BdcReportService._pdf_cell(row.phone, body_cell),
                BdcReportService._pdf_cell(row.stage, body_cell),
                BdcReportService._pdf_cell(row.dealership, body_cell),
                BdcReportService._pdf_cell(row.bdc_agent, body_cell),
                BdcReportService._pdf_cell(row.salesperson, body_cell),
                BdcReportService._pdf_cell(row.latest_appt_status, body_cell),
                Paragraph(appt_display, body_cell),
                Paragraph(trust_display, body_cell),
                qr_cell,
                BdcReportService._pdf_cell("Yes" if row.guest_auto_generated else "", body_cell_center),
            ])

        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        style_commands: List[Any] = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 7.5),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("ALIGN", (qr_col_idx, 1), (qr_col_idx, -1), "CENTER"),
            ("ALIGN", (-1, 1), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E2E8F0")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ]
        for row_i in highlight_rows:
            style_commands.append(
                ("BACKGROUND", (0, row_i), (-1, row_i), colors.HexColor("#FEF3C7"))
            )
        table.setStyle(TableStyle(style_commands))
        elements.append(table)

        def _draw_page_footer(canvas: Any, doc_obj: Any) -> None:
            canvas.saveState()
            canvas.setStrokeColor(colors.HexColor("#CBD5E1"))
            canvas.setLineWidth(0.5)
            canvas.line(
                side_margin,
                10 * mm,
                page_w - side_margin,
                10 * mm,
            )
            canvas.setFont("Helvetica-Bold", 7)
            canvas.setFillColor(colors.HexColor("#334155"))
            canvas.drawString(side_margin, 6 * mm, f"{settings.app_name}")
            canvas.setFont("Helvetica", 7)
            canvas.setFillColor(colors.HexColor("#64748B"))
            brand_w = canvas.stringWidth(f"{settings.app_name}", "Helvetica-Bold", 7)
            canvas.drawString(side_margin + brand_w, 6 * mm, "  ·  BDC Export Report")
            if meta:
                canvas.drawCentredString(
                    page_w / 2,
                    6 * mm,
                    f"{meta.total_leads} leads · {meta.scope_label}",
                )
            canvas.drawRightString(
                page_w - side_margin,
                6 * mm,
                f"Page {canvas.getPageNumber()}",
            )
            canvas.restoreState()

        doc.build(elements, onFirstPage=_draw_page_footer, onLaterPages=_draw_page_footer)
        return buf.getvalue()

    @staticmethod
    def row_to_dict(row: BdcReportRow) -> Dict[str, Any]:
        return {
            "lead_id": row.lead_id,
            "full_name": row.full_name,
            "email": row.email,
            "phone": row.phone,
            "stage": row.stage,
            "source": row.source,
            "lead_created": row.lead_created,
            "dealership": row.dealership,
            "bdc_agent": row.bdc_agent,
            "salesperson": row.salesperson,
            "is_active": row.is_active,
            "converted_at": row.converted_at,
            "latest_appt_status": row.latest_appt_status,
            "latest_appt_date": row.latest_appt_date,
            "appt_count": row.appt_count,
            "showroom_check_in": row.showroom_check_in,
            "lead_trust_score": row.lead_trust_score,
            "guest_trust_score": row.guest_trust_score,
            "guest_qr_url": row.guest_qr_url,
            "guest_auto_generated": row.guest_auto_generated,
        }
